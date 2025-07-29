import pandas as pd
import openpyxl
from datetime import datetime

def fix_excel_answers():
    excel_file = '와석초_개선된QA_데이터_링크포함_20250729_114733.xlsx'
    
    try:
        # 엑셀 파일 읽기
        workbook = openpyxl.load_workbook(excel_file)
        
        print(f"엑셀 파일 '{excel_file}' 읽기 성공!")
        print(f"시트 개수: {len(workbook.sheetnames)}")
        
        # 제외할 시트들
        exclude_sheets = ['강화된_QA_데이터', '원본_QA_데이터', '강화_통계', '크롤링_요약']
        
        for sheet_name in workbook.sheetnames:
            if any(exclude in sheet_name for exclude in exclude_sheets):
                print(f"시트 '{sheet_name}' 제외")
                continue
                
            print(f"\n=== 시트: {sheet_name} 처리 중 ===")
            
            worksheet = workbook[sheet_name]
            
            # 데이터 범위 찾기
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            print(f"행 수: {max_row}, 열 수: {max_col}")
            
            # 열 인덱스 찾기
            header_row = 1
            question_col = None
            answer_col = None
            link_col = None
            
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=header_row, column=col).value
                if cell_value == '질문':
                    question_col = col
                elif cell_value == '답변':
                    answer_col = col
                elif cell_value == '링크':
                    link_col = col
            
            if not all([question_col, answer_col, link_col]):
                print(f"필수 열을 찾을 수 없음: 질문={question_col}, 답변={answer_col}, 링크={link_col}")
                continue
            
            # 각 행 처리
            modified_count = 0
            for row in range(2, max_row + 1):  # 헤더 제외
                question = worksheet.cell(row=row, column=question_col).value
                answer = worksheet.cell(row=row, column=answer_col).value
                link = worksheet.cell(row=row, column=link_col).value
                
                # 빈 행 건너뛰기
                if not question or str(question).strip() == '':
                    continue
                
                # 링크가 있으면 답변에 포함
                if link and str(link).strip() != '' and str(link).strip() != 'nan':
                    # 답변 + 링크 형태로 합치기
                    new_answer = f"{answer}\n\n{link}"
                    worksheet.cell(row=row, column=answer_col).value = new_answer
                    modified_count += 1
                    print(f"수정: {str(question)[:30]}...")
            
            print(f"수정된 행: {modified_count}개")
        
        # 백업 파일 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = f'와석초_개선된QA_데이터_백업_{timestamp}.xlsx'
        workbook.save(backup_file)
        print(f"\n백업 파일 생성: {backup_file}")
        
        # 새 파일명으로 저장
        new_file = f'와석초_개선된QA_데이터_답변링크합침_{timestamp}.xlsx'
        workbook.save(new_file)
        print(f"새 파일 저장 완료: {new_file}")
        
        return True
        
    except Exception as e:
        print(f"오류 발생: {e}")
        return False

if __name__ == "__main__":
    success = fix_excel_answers()
    if success:
        print("\n✅ 엑셀 답변+링크 합침 완료!")
    else:
        print("\n❌ 처리 실패") 