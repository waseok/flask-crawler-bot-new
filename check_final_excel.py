import pandas as pd
import openpyxl

def check_final_excel():
    excel_file = '와석초_답변링크합침_최종.xlsx'
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        
        print(f"최종 엑셀 파일 '{excel_file}' 확인 중...")
        
        # 급식정보 시트 확인
        if '급식정보' in workbook.sheetnames:
            worksheet = workbook['급식정보']
            print("\n=== 급식정보 시트 확인 ===")
            
            # 첫 번째 질문의 답변 확인
            for row in range(2, 4):  # 2-3행 확인
                question = worksheet.cell(row=row, column=3).value  # 질문 열
                answer = worksheet.cell(row=row, column=4).value    # 답변 열
                link = worksheet.cell(row=row, column=5).value      # 링크 열
                
                if question:
                    print(f"\n질문: {question}")
                    print(f"답변: {answer}")
                    print(f"링크 열: {link}")
                    print("-" * 50)
        
        # 방과후 시트 확인
        if '방과후' in workbook.sheetnames:
            worksheet = workbook['방과후']
            print("\n=== 방과후 시트 확인 ===")
            
            # 첫 번째 질문의 답변 확인
            for row in range(2, 4):  # 2-3행 확인
                question = worksheet.cell(row=row, column=3).value  # 질문 열
                answer = worksheet.cell(row=row, column=4).value    # 답변 열
                link = worksheet.cell(row=row, column=5).value      # 링크 열
                
                if question:
                    print(f"\n질문: {question}")
                    print(f"답변: {answer}")
                    print(f"링크 열: {link}")
                    print("-" * 50)
        
        return True
        
    except Exception as e:
        print(f"오류 발생: {e}")
        return False

if __name__ == "__main__":
    success = check_final_excel()
    if success:
        print("\n✅ 최종 엑셀 확인 완료!")
    else:
        print("\n❌ 확인 실패") 